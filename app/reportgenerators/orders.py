import json
import pandas as pd
import openpyxl
import shutil
import re
from openpyxl.styles import numbers

from configs.companynameenum import CompanyName
from configs.dirconfigenum import DirPath
from configs.reporttypeenum import ReportType
from utils.commonutil import get_company, get_price, get_template 


def generate_orders():
    print(f'o(^^o)')
    print(f'(o^^)o ~~ Generating order summary report...')
    print(f'o(^^o)')

    avail_companies = [CompanyName.SEA, CompanyName.LOR, CompanyName.ULT]

    def read_config(file_path):
        with open(file_path, 'r') as file:
            config = json.load(file)
        return config

    file_config = read_config(DirPath.ORDER_FILE.value)
    lookup_config = read_config(DirPath.ORDER_LOOKUP.value)

    file_input = file_config['myob']['file']
    input_delimiter = file_config['myob']['delimiter']
    input_header = file_config['myob']['header']
    sheetname = file_config['excel']['orders']['sheetname']
    fhc_code = file_config['excel']['orders']['fhc_code']
    company_col = file_config['excel']['orders']['company_col']
    title_web = file_config['excel']['orders']['title_web']
    title_contact = file_config['excel']['orders']['title_contact']
    lookup = lookup_config['lookup']


    def get_period(date):
        if date.day <= 15:
            return f'1-15 {date.strftime("%B %Y")}'
        else:
            last_day_of_month = date.replace(day=1, month=date.month+1) - pd.DateOffset(days=1)
            return f'16-{last_day_of_month.day} {date.strftime("%B %Y")}'


    def get_store(lastname, code):
        match = re.search(lookup[code]["store_pattern"], lastname)
        if match:
            return match.group(1)

        return ""


    df = pd.read_table(file_input,delimiter=input_delimiter,header=input_header)


    # Data prep
    company_df = df.copy()
    company_df = company_df[[company_col, "Invoice No.", "Date", "Item Number", "Description", "Quantity", "Price"]]
    company_df.fillna("NaN", inplace=True)


    # Metadata
    [code, company] = get_company(company_df.iloc[0][company_col], avail_companies)
    template = get_template(ReportType.ORDERS.value, code)
    sheet = template[sheetname]


    # Date format, period, stores
    company_df["Date"] = pd.to_datetime(company_df["Date"], format="%d/%m/%Y")
    company_df["Period"] = company_df["Date"].apply(get_period)
    company_df["Store"] = company_df["Co./Last Name"].apply(lambda row: get_store(row, code))


    # Populate template
    pivot_df = company_df.pivot_table(index='Item Number', columns=['Invoice No.', 'Store', 'Date'], values='Quantity', fill_value=0)
    pivot_df = pivot_df.reindex(sorted(pivot_df.columns, key=lambda x: x[2]), axis=1) # Sort cols by date asc

    # display(pivot_df)

    template_item_codes = set(sheet[f'{lookup[code]["template_code"]["col"]}{row}'].value for row in range(lookup[code]["template_code"]["start_row"], lookup[code]["template_code"]["end_row"] + 1))
    new_item_codes = set(company_df['Item Number'].unique()) - template_item_codes

    # -- Append new items found to template
    for idx, new_item_code in enumerate(new_item_codes, start=lookup[code]["template_code"]["end_row"] + 1):
        sheet[f'{lookup[code]["template_code"]["col"]}{idx}'] = new_item_code
        sheet[f'{lookup[code]["template_item"]["col"]}{idx}'] = company_df[company_df['Item Number'] == new_item_code]['Description'].values[0]
        sheet[f'{lookup[code]["template_price"]["col"]}{idx}'] = get_price(company_df[company_df['Item Number'] == new_item_code]['Price'].values[0])

    # -- Write dates invoice, row, stores rows
    for col_num, (invoice, store, date) in enumerate(pivot_df.columns, start=lookup[code]["template_date"]["start_row"]):
        sheet.cell(row=lookup[code]["template_date"]["start_row"], column=col_num, value=date)
        sheet.cell(row=lookup[code]["template_date"]["start_row"] + 1, column=col_num, value=store)
        sheet.cell(row=lookup[code]["template_extra"]["end_row"] + 5, column=col_num, value=invoice)

    # -- Fill in quantities
    for row_num, code_row in enumerate(range(lookup[code]["template_code"]["start_row"], lookup[code]["template_extra"]["end_row"] + 1), start=lookup[code]["template_code"]["start_row"]):
        item_code_val = sheet[f'{lookup[code]["template_code"]["col"]}{code_row}'].value
        
        if item_code_val in pivot_df.index:
            for col_num, (invoice, store, date) in enumerate(pivot_df.columns, start=lookup[code]["template_date"]["start_row"]):
                if item_code_val == fhc_code:
                    # Special case for FHC
                    price_entry = company_df[(company_df['Item Number'] == fhc_code) & (company_df['Store'] == store) & (company_df['Date'] == date)]
                    if not price_entry.empty:
                        price_str = price_entry['Price'].values[0]
                        price = get_price(price_str)
                    else:
                        price = ""

                    cell_val = price
                else:
                    # Get quantity
                    qty = pivot_df.loc[item_code_val, (invoice, store, date)]
                    cell_val = "" if qty == 0 else qty
                
                sheet.cell(row=row_num, column=col_num, value=cell_val)


    # Output
    period = company_df["Period"].iloc[0]
    period_str = f'for the Period of {period}'
    sheet.cell(row=1, column=2, value=f'{company.upper()} - Summary of Orders')
    sheet.cell(row=2, column=2, value=period_str)
    sheet.cell(row=3, column=2, value=title_web)
    sheet.cell(row=4, column=2, value=title_contact)

    new_filename = f'{company.title().replace(" ", "_")}_All_Stores_{period.replace(" ", "_")}.xlsx'
    output_path = f'{DirPath.OUT_ORDER.value}{new_filename}'
    template.save(output_path)

    print(f'Data saved to {output_path}')
